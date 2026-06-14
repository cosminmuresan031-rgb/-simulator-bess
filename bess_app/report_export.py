from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any, Dict, List, Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


MONTHS = ["Ian", "Feb", "Mar", "Apr", "Mai", "Iun", "Iul", "Aug", "Sep", "Oct", "Nov", "Dec"]
NAVY = "06152C"
GREEN = "4F9B27"
GREEN_DARK = "2D711B"
GREEN_LIGHT = "E8F4EB"
BLUE_LIGHT = "D7E8F5"
RED_LIGHT = "F7DEDE"
TEXT = "0A1D32"
WHITE = "FFFFFF"
LINE = "D9E3EB"


def as_float(value: Any) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def safe_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    return "'" + text if text.startswith("=") else text


def sanitize_filename_part(value: Any, fallback: str = "raport") -> str:
    text = str(value or fallback).strip()
    text = re.sub(r"[^A-Za-z0-9_-]+", "_", text)
    return text.strip("_") or fallback


def safe_sheet_title(value: Any, fallback: str, existing: set[str]) -> str:
    title = re.sub(r"[\[\]\:\*\?\/\\]+", " ", str(value or fallback)).strip() or fallback
    title = title[:31]
    base = title
    suffix = 2
    while title in existing:
        marker = f" {suffix}"
        title = f"{base[:31 - len(marker)]}{marker}"
        suffix += 1
    existing.add(title)
    return title


def style_title(cell) -> None:
    cell.font = Font(bold=True, size=18, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_section(cell) -> None:
    cell.font = Font(bold=True, size=12, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_header(row) -> None:
    for cell in row:
        cell.font = Font(bold=True, color=TEXT)
        cell.fill = PatternFill("solid", fgColor=BLUE_LIGHT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color=LINE))


def style_table(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    thin = Side(style="thin", color=LINE)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(bottom=thin, right=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"


def set_widths(ws, widths: Dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_kpi(ws, row: int, col: int, label: str, value: Any) -> None:
    label_cell = ws.cell(row=row, column=col, value=safe_text(label))
    value_cell = ws.cell(row=row + 1, column=col, value=value)
    label_cell.font = Font(color="526B72", size=10)
    value_cell.font = Font(bold=True, color=TEXT, size=14)
    label_cell.fill = PatternFill("solid", fgColor=GREEN_LIGHT)
    value_cell.fill = PatternFill("solid", fgColor=GREEN_LIGHT)
    label_cell.alignment = Alignment(horizontal="center")
    value_cell.alignment = Alignment(horizontal="center")
    if isinstance(value, (int, float)):
        value_cell.number_format = "#,##0.00"


def write_rows(ws, start_row: int, headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> int:
    for col, header in enumerate(headers, start=1):
        ws.cell(start_row, col, safe_text(header))
    style_header(ws[start_row])
    for row_index, row_values in enumerate(rows, start=start_row + 1):
        for col, value in enumerate(row_values, start=1):
            ws.cell(row_index, col, safe_text(value) if isinstance(value, str) else value)
    max_row = start_row + len(rows)
    style_table(ws, start_row, max_row, 1, len(headers))
    return max_row


def month_label(row: Dict[str, Any], index: int) -> str:
    return str(row.get("month") or (MONTHS[index] if index < len(MONTHS) else index + 1))


def bess_action(row: Dict[str, Any]) -> str:
    if as_float(row.get("charge_mwh")) > 0:
        return "Incarcare"
    if as_float(row.get("discharge_mwh")) > 0:
        return "Descarcare"
    return "Standby"


def create_dashboard_sheet(wb: Workbook, dashboard: Dict[str, Any], project_name: str) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A12"
    set_widths(
        ws,
        {"A": 16, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16, "I": 4, "J": 16, "K": 16, "L": 16, "M": 16},
    )

    ws.merge_cells("A1:H1")
    ws["A1"] = "SIMULATOR BESS - DASHBOARD"
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 28

    metrics = dashboard.get("metrics") or {}
    annual = dashboard.get("annual") or {}
    battery = dashboard.get("battery") or {}
    ws["A3"] = "Proiect"
    ws["B3"] = safe_text(project_name or "-")
    ws["A4"] = "Scenariu selectat"
    ws["B4"] = safe_text(dashboard.get("scenarioName") or "-")
    ws["A5"] = "Generat la"
    ws["B5"] = safe_text(dashboard.get("generatedAt") or datetime.now().isoformat(timespec="seconds"))
    ws["D3"] = "Baterie MW AC"
    ws["E3"] = as_float(battery.get("acMw"))
    ws["D4"] = "Stocare MWh"
    ws["E4"] = as_float(battery.get("storageMwh"))
    ws["D5"] = "Ore analizate"
    ws["E5"] = as_float(metrics.get("hours"))

    for cell in ["A3", "A4", "A5", "D3", "D4", "D5"]:
        ws[cell].font = Font(bold=True, color=TEXT)
    for cell in ["E3", "E4", "E5"]:
        ws[cell].number_format = "#,##0.00"

    add_kpi(ws, 7, 1, "Fara BESS", as_float(annual.get("arithmetic_avg_monthly_without_bess_lei_mwh")))
    add_kpi(ws, 7, 3, "Cu BESS", as_float(annual.get("arithmetic_avg_monthly_with_bess_lei_mwh")))
    add_kpi(ws, 7, 5, "Reducere", as_float(annual.get("reduction_lei_mwh")))
    add_kpi(ws, 7, 7, "Consum MWh", as_float(annual.get("consumption_mwh") or metrics.get("totalConsumptionMwh")))

    scenario_rows = [
        [
            safe_text(row.get("name") or "-"),
            as_float(row.get("acMw")),
            as_float(row.get("storageMwh")),
            as_float(row.get("costWithoutBessLei")),
            as_float(row.get("costWithBessLei")),
            as_float(row.get("priceWithoutBessLeiMwh")),
            as_float(row.get("priceWithBessLeiMwh")),
            as_float(row.get("reductionLeiMwh")),
        ]
        for row in (dashboard.get("scenarios") or [])
    ]

    ws.merge_cells("A11:H11")
    ws["A11"] = "Scenarii si costuri"
    style_section(ws["A11"])
    scenario_end_row = write_rows(
        ws,
        12,
        ["Scenariu", "MW AC", "MWh", "Cost fara BESS", "Cost cu BESS", "Pret fara BESS", "Pret cu BESS", "Reducere"],
        scenario_rows,
    )

    if scenario_rows:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Comparator scenarii - pret mediu anual"
        chart.y_axis.title = "lei/MWh"
        chart.x_axis.title = "Scenariu"
        data = Reference(ws, min_col=6, max_col=7, min_row=12, max_row=scenario_end_row)
        cats = Reference(ws, min_col=1, min_row=13, max_row=scenario_end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 7
        chart.width = 16
        ws.add_chart(chart, "J12")

    monthly_section_row = scenario_end_row + 3
    ws.merge_cells(start_row=monthly_section_row, start_column=1, end_row=monthly_section_row, end_column=8)
    ws.cell(monthly_section_row, 1, "Preturi medii lunare")
    style_section(ws.cell(monthly_section_row, 1))

    monthly = dashboard.get("monthly") or []
    rows = [
        [
            month_label(row, index),
            as_float(row.get("price_without_bess_lei_mwh")),
            as_float(row.get("price_with_bess_lei_mwh")),
            as_float(row.get("reduction_lei_mwh")),
            as_float(row.get("charge_count")),
            as_float(row.get("discharge_count")),
        ]
        for index, row in enumerate(monthly)
    ]
    end_row = write_rows(
        ws,
        monthly_section_row + 1,
        ["Luna", "Fara BESS", "Cu BESS", "Reducere", "Incarcari", "Descarcari"],
        rows,
    )

    if rows:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Fara BESS vs Cu BESS"
        chart.y_axis.title = "lei/MWh"
        chart.x_axis.title = "Luna"
        data = Reference(ws, min_col=2, max_col=3, min_row=monthly_section_row + 1, max_row=end_row)
        cats = Reference(ws, min_col=1, min_row=monthly_section_row + 2, max_row=end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 7
        chart.width = 16
        ws.add_chart(chart, f"H{monthly_section_row + 1}")


def create_financial_sheet(wb: Workbook, summary: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Sinteza financiara")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A10"
    set_widths(ws, {"A": 24, "B": 18, "C": 18, "D": 18, "E": 18, "F": 20, "G": 34, "H": 4, "I": 18, "J": 18, "K": 18})

    assumptions = summary.get("assumptions") or {}
    annual = summary.get("annual") or {}
    battery = summary.get("battery") or {}
    interpretation = summary.get("interpretation") or {}
    monthly = summary.get("monthly") or []
    supply_component = as_float(annual.get("supply_component_lei_mwh"))
    gross_diff = as_float(annual.get("difference_vs_contract_lei_mwh") or annual.get("contract_vs_pzu_with_bess_lei_mwh"))
    final_result_value = annual.get("final_result_lei_mwh")
    final_result = as_float(final_result_value) if final_result_value not in (None, "") else gross_diff - supply_component

    ws.merge_cells("A1:G1")
    ws["A1"] = "SINTEZA FINANCIARA"
    style_title(ws["A1"])
    ws["A3"] = "Client"
    ws["B3"] = safe_text(assumptions.get("client") or "Client")
    ws["A4"] = "Scenariu"
    ws["B4"] = safe_text(summary.get("scenarioName") or "-")
    ws["A5"] = "Capacitate baterie"
    ws["B5"] = f"{as_float(battery.get('acMw')):.3f} MW AC / {as_float(battery.get('storageMwh')):.3f} MWh"
    ws["A6"] = "Consum zi / noapte"
    ws["B6"] = f"{as_float(interpretation.get('dayPct')):.1f}% / {as_float(interpretation.get('nightPct')):.1f}%"

    indicator_rows = [
        ["Consum total analizat [MWh]", as_float(annual.get("consumption_mwh")), "Volumul anual pe care se aplica optimizarea."],
        ["Pret mediu ponderat cu BESS", as_float(annual.get("weighted_price_with_bess_lei_mwh")), "Reperul de cost cu stocare."],
        ["Pret contractual furnizare 2025", as_float(annual.get("contract_price_lei_mwh")), "Nivelul de comparatie comerciala."],
        ["Componenta furnizare PZU", supply_component, "Componenta editabila scazuta din rezultatul anual."],
        ["Impact indicativ [lei/an]", as_float(annual.get("impact_lei_year")), "Consum total x rezultat net."],
        ["Termen recuperare [ani]", as_float(annual.get("payback_years")), "Calculat pe impact anual estimat."],
    ]
    ws.merge_cells("A8:G8")
    ws["A8"] = "Indicatori"
    style_section(ws["A8"])
    write_rows(ws, 9, ["Indicator", "Rezultat", "Interpretare"], indicator_rows)

    start = 17
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=7)
    ws.cell(start, 1, "Preturi medii lunare")
    style_section(ws.cell(start, 1))
    monthly_rows = [
        [
            month_label(row, index),
            as_float(row.get("contract_price_lei_mwh")),
            as_float(row.get("price_with_bess_lei_mwh")),
            as_float(row.get("difference_vs_contract_lei_mwh")),
            safe_text(row.get("observation") or ""),
        ]
        for index, row in enumerate(monthly)
    ]
    end_month = write_rows(
        ws,
        start + 1,
        [
            "Luna",
            "Pret contract 2025",
            "PZU cu BESS",
            "Pret contract 2025 vs PZU cu BESS",
            "Observatie",
        ],
        monthly_rows,
    )
    if monthly_rows:
        ws.conditional_formatting.add(
            f"D{start + 2}:D{end_month}",
            DataBarRule(start_type="min", end_type="max", color="70AD47", showValue=True),
        )

    if monthly_rows:
        client = safe_text(assumptions.get("client") or "Client")
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = f"{client} - {summary.get('scenarioName') or '-'} | preturi lunare"
        chart.y_axis.title = "lei/MWh"
        chart.x_axis.title = "Luna"
        price_values = [
            as_float(row[1]) for row in monthly_rows if len(row) > 1
        ] + [
            as_float(row[2]) for row in monthly_rows if len(row) > 2
        ]
        if price_values:
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = max(price_values) * 1.15
        data = Reference(ws, min_col=2, max_col=3, min_row=start + 1, max_row=end_month)
        cats = Reference(ws, min_col=1, min_row=start + 2, max_row=end_month)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        if len(chart.series) >= 2:
            chart.series[0].graphicalProperties.solidFill = NAVY
            chart.series[0].graphicalProperties.line.solidFill = NAVY
            chart.series[1].graphicalProperties.solidFill = GREEN
            chart.series[1].graphicalProperties.line.solidFill = GREEN_DARK
        chart.height = 7
        chart.width = 16
        ws.add_chart(chart, "I18")

    annual_start = end_month + 3
    ws.merge_cells(start_row=annual_start, start_column=1, end_row=annual_start, end_column=7)
    ws.cell(annual_start, 1, "Sinteza anuala")
    style_section(ws.cell(annual_start, 1))
    annual_rows = [
        [
            "Medie lunara rezultate",
            as_float(annual.get("contract_price_lei_mwh")),
            as_float(annual.get("arithmetic_avg_monthly_with_bess_lei_mwh")),
            gross_diff,
            supply_component,
            final_result,
            "Medii simple ale celor 12 luni din simulare",
        ],
        ["Impact indicativ [lei/an]", "", "", "", "", as_float(annual.get("impact_lei_year")), "Consum total x rezultat net"],
        ["Impact indicativ [euro/an]", "", "", "", "", as_float(annual.get("impact_eur_year")), "Conversie la cursul introdus"],
        ["Valoare investitie [euro]", "", "", "", "", as_float(annual.get("investment_eur")), "Ipoteza introdusa"],
        ["Termen de recuperare investitie ani", "", "", "", "", as_float(annual.get("payback_years")), "Investitie / impact anual euro"],
    ]
    write_rows(
        ws,
        annual_start + 1,
        ["Sinteza anuala", "Pret contract 2025", "PZU cu BESS", "Contract vs PZU cu BESS", "Componenta furnizare", "Rezultate", "Comentariu"],
        annual_rows,
    )


def create_scenario_sheet(wb: Workbook, scenario: Dict[str, Any], title: str = "Scenariu selectat") -> None:
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A9"
    widths = {
        "A": 20,
        "B": 14,
        "C": 12,
        "D": 14,
        "E": 14,
        "F": 14,
        "G": 14,
        "H": 14,
        "I": 14,
        "J": 14,
        "K": 14,
        "L": 14,
        "M": 16,
        "N": 16,
        "O": 16,
        "P": 16,
    }
    set_widths(ws, widths)

    ws.merge_cells("A1:P1")
    ws["A1"] = "SCENARIU SELECTAT - COMPORTAMENT INCARCARE / DESCARCARE"
    style_title(ws["A1"])
    ws["A3"] = "Scenariu"
    ws["B3"] = safe_text(scenario.get("name") or "-")
    ws["A4"] = "MW AC"
    ws["B4"] = as_float(scenario.get("ac_mw"))
    ws["A5"] = "Stocare MWh"
    ws["B5"] = as_float(scenario.get("storage_mwh"))
    ws["D3"] = "Ore incarcare candidat"
    ws["E3"] = as_float(scenario.get("charge_candidate_hours"))
    ws["D4"] = "Ore descarcare candidat"
    ws["E4"] = as_float(scenario.get("discharge_candidate_hours"))

    annual = (scenario.get("summary") or {}).get("annual") or {}
    ws["G3"] = "Pret mediu fara BESS"
    ws["H3"] = as_float(annual.get("arithmetic_avg_monthly_without_bess_lei_mwh"))
    ws["G4"] = "Pret mediu cu BESS"
    ws["H4"] = as_float(annual.get("arithmetic_avg_monthly_with_bess_lei_mwh"))
    ws["G5"] = "Reducere"
    ws["H5"] = as_float(annual.get("reduction_lei_mwh"))

    for cell in ["A3", "A4", "A5", "D3", "D4", "G3", "G4", "G5"]:
        ws[cell].font = Font(bold=True, color=TEXT)
    for cell in ["B4", "B5", "E3", "E4", "H3", "H4", "H5"]:
        ws[cell].number_format = "#,##0.00"

    headers = [
        "Timestamp",
        "Consum MWh",
        "PV MWh",
        "Pret lei/MWh",
        "Net load MWh",
        "SOC start MWh",
        "Incarcare MWh",
        "Incarcare retea MWh",
        "Incarcare PV MWh",
        "Descarcare MWh",
        "Actiune BESS",
        "SOC final MWh",
        "Achizitie neta MWh",
        "Cost fara BESS lei",
        "Cost cu BESS lei",
        "Economie lei",
    ]
    hourly = scenario.get("hourly") or []
    rows = [
        [
            safe_text(row.get("timestamp") or ""),
            as_float(row.get("consumption_mwh")),
            as_float(row.get("pv_mwh")),
            as_float(row.get("price_lei_mwh")),
            as_float(row.get("net_load_mwh")),
            as_float(row.get("soc_start_mwh")),
            as_float(row.get("charge_mwh")),
            as_float(row.get("grid_charge_mwh")),
            as_float(row.get("pv_charge_mwh")),
            as_float(row.get("discharge_mwh")),
            bess_action(row),
            as_float(row.get("soc_final_mwh")),
            as_float(row.get("net_purchase_mwh")),
            as_float(row.get("cost_without_bess_lei")),
            as_float(row.get("cost_with_bess_lei")),
            as_float(row.get("saving_lei")),
        ]
        for row in hourly
    ]
    end_row = write_rows(ws, 8, headers, rows)
    ws.auto_filter.ref = f"A8:P{end_row}"
    for row_index in range(9, end_row + 1):
        action = str(ws.cell(row_index, 11).value or "")
        if action == "Incarcare":
            fill = PatternFill("solid", fgColor=GREEN_LIGHT)
        elif action == "Descarcare":
            fill = PatternFill("solid", fgColor=RED_LIGHT)
        else:
            continue
        ws.cell(row_index, 11).fill = fill
        ws.cell(row_index, 11).font = Font(bold=True, color=TEXT)


def create_procurement_dna_sheet(wb: Workbook, dna: Dict[str, Any]) -> None:
    if not dna:
        return
    ws = wb.create_sheet("Procurement DNA")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A12"
    set_widths(ws, {"A": 28, "B": 20, "C": 20, "D": 20, "E": 20, "F": 20, "G": 22, "H": 4, "I": 18, "J": 18, "K": 18})

    ws.merge_cells("A1:G1")
    ws["A1"] = "PROCUREMENT DNA"
    style_title(ws["A1"])
    ws.row_dimensions[1].height = 28

    load_dna = dna.get("loadDna") or {}
    cost_dna = dna.get("costDna") or {}
    summary_rows = [
        ["Procurement DNA", safe_text(dna.get("finalOutput") or "-")],
        ["Confidence Score", as_float(dna.get("confidenceScore"))],
        ["Risk Gauge", as_float(dna.get("riskScore"))],
        ["Recomandare contract", safe_text(dna.get("contractRecommendation") or "-")],
        ["Risk Level", safe_text(dna.get("riskLevel") or "-")],
        ["Ore valide consum", as_float(dna.get("validConsumptionHours"))],
    ]
    ws.merge_cells("A3:G3")
    ws["A3"] = "Executive DNA Summary"
    style_section(ws["A3"])
    write_rows(ws, 4, ["Indicator", "Rezultat"], summary_rows)

    risk_rows = [
        ["Forecastability Score", as_float(dna.get("forecastabilityScore"))],
        ["Seasonality Score", as_float(dna.get("seasonalityScore"))],
        ["Volatility Score", as_float(dna.get("volatilityScore"))],
        ["Peak Risk Score", as_float(dna.get("peakRiskScore"))],
        ["Risk Gauge Total", as_float(dna.get("riskScore"))],
    ]
    ws.merge_cells("D4:G4")
    ws["D4"] = "Risk DNA"
    style_section(ws["D4"])
    for row_offset, values in enumerate(risk_rows, start=5):
        ws.cell(row_offset, 4, values[0])
        ws.cell(row_offset, 5, values[1])
    style_table(ws, 5, 4 + len(risk_rows), 4, 5)

    load_rows = [
        ["Baseload MW", as_float(load_dna.get("baseloadMw"))],
        ["Midload MW", as_float(load_dna.get("midloadMw"))],
        ["Peakload MW", as_float(load_dna.get("peakloadMw"))],
        ["Peak/Base Ratio", as_float(load_dna.get("peakBaseRatio"))],
        ["Interpretare", safe_text(load_dna.get("label") or "-")],
    ]
    ws.merge_cells("A13:G13")
    ws["A13"] = "Load DNA"
    style_section(ws["A13"])
    load_end = write_rows(ws, 14, ["Indicator", "Rezultat"], load_rows)

    cost_rows = [
        ["Total Cost PZU", as_float(cost_dna.get("totalCost"))],
        ["Cost Concentration Index", as_float(cost_dna.get("concentrationIndex"))],
        ["Top 100 Hours Share", as_float(cost_dna.get("top100Share"))],
        ["Expensive Hours Cost Share", as_float(cost_dna.get("expensiveCostShare"))],
        ["Cheap Hours Cost Share", as_float(cost_dna.get("cheapCostShare"))],
    ]
    ws.merge_cells("D14:G14")
    ws["D14"] = "Cost DNA"
    style_section(ws["D14"])
    for row_offset, values in enumerate(cost_rows, start=15):
        ws.cell(row_offset, 4, values[0])
        ws.cell(row_offset, 5, values[1])
    style_table(ws, 15, 14 + len(cost_rows), 4, 5)

    report_row = load_end + 3
    ws.merge_cells(start_row=report_row, start_column=1, end_row=report_row, end_column=7)
    ws.cell(report_row, 1, "Interpretare automata profesionala")
    style_section(ws.cell(report_row, 1))
    ws.merge_cells(start_row=report_row + 1, start_column=1, end_row=report_row + 4, end_column=7)
    report_cell = ws.cell(report_row + 1, 1, safe_text(dna.get("reportText") or ""))
    report_cell.alignment = Alignment(vertical="top", wrap_text=True)

    ldc_rows = [
        [as_float(row.get("hour")), as_float(row.get("value"))]
        for row in (load_dna.get("points") or [])
    ]
    ldc_start = report_row + 7
    ws.merge_cells(start_row=ldc_start, start_column=1, end_row=ldc_start, end_column=3)
    ws.cell(ldc_start, 1, "Load Duration Curve")
    style_section(ws.cell(ldc_start, 1))
    ldc_end = write_rows(ws, ldc_start + 1, ["Ora sortata", "Consum MW"], ldc_rows)
    if ldc_rows:
        chart = LineChart()
        chart.style = 10
        chart.title = "Load Duration Curve"
        chart.y_axis.title = "MW"
        chart.x_axis.title = "Ore sortate"
        data = Reference(ws, min_col=2, min_row=ldc_start + 1, max_row=ldc_end)
        cats = Reference(ws, min_col=1, min_row=ldc_start + 2, max_row=ldc_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 7
        chart.width = 14
        ws.add_chart(chart, f"D{ldc_start + 1}")

    top_rows = [
        [
            safe_text(row.get("timestamp") or ""),
            as_float(row.get("consumption")),
            as_float(row.get("price")),
            as_float(row.get("cost")),
        ]
        for row in (cost_dna.get("top10") or [])
    ]
    top_start = ldc_end + 3
    ws.merge_cells(start_row=top_start, start_column=1, end_row=top_start, end_column=4)
    ws.cell(top_start, 1, "Top 10 Cost Hours")
    style_section(ws.cell(top_start, 1))
    top_end = write_rows(ws, top_start + 1, ["Ora", "Consum MWh", "Pret PZU", "Cost lei"], top_rows)
    if top_rows:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Top 10 Cost Hours"
        chart.x_axis.title = "lei"
        data = Reference(ws, min_col=4, min_row=top_start + 1, max_row=top_end)
        cats = Reference(ws, min_col=1, min_row=top_start + 2, max_row=top_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 7
        chart.width = 14
        ws.add_chart(chart, f"F{top_start + 1}")

    distribution_rows = [[safe_text(row.get("label") or ""), as_float(row.get("value"))] for row in (cost_dna.get("distribution") or [])]
    distribution_start = top_end + 3
    ws.merge_cells(start_row=distribution_start, start_column=1, end_row=distribution_start, end_column=3)
    ws.cell(distribution_start, 1, "Cost Distribution")
    style_section(ws.cell(distribution_start, 1))
    distribution_end = write_rows(ws, distribution_start + 1, ["Categorie", "Cost lei"], distribution_rows)
    if distribution_rows:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Cost Distribution"
        chart.y_axis.title = "lei"
        data = Reference(ws, min_col=2, min_row=distribution_start + 1, max_row=distribution_end)
        cats = Reference(ws, min_col=1, min_row=distribution_start + 2, max_row=distribution_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 7
        chart.width = 14
        ws.add_chart(chart, f"D{distribution_start + 1}")


def assert_no_formulas(wb: Workbook) -> None:
    formulas: List[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append(f"{ws.title}!{cell.coordinate}")
                    if len(formulas) >= 10:
                        break
            if len(formulas) >= 10:
                break
        if len(formulas) >= 10:
            break
    if formulas:
        raise ValueError("Raportul contine formule in: " + ", ".join(formulas))


def generate_complete_report_xlsx(payload: Dict[str, Any]) -> bytes:
    dashboard = payload.get("dashboard") or {}
    summary = payload.get("financialSummary") or {}
    scenario = payload.get("scenario") or {}
    procurement_dna = payload.get("procurementDna") or {}
    project_name = str(payload.get("projectName") or "Proiect BESS")

    if not dashboard:
        raise ValueError("Dashboard lipsa pentru export.")
    if not summary:
        raise ValueError("Sinteza financiara lipsa pentru export.")
    if not scenario:
        raise ValueError("Scenariul selectat lipseste pentru export.")

    wb = Workbook()
    default = wb.active
    wb.remove(default)
    wb.properties.creator = "Simulator BESS"
    wb.properties.title = "Raport BESS valori inghetate"

    create_dashboard_sheet(wb, dashboard, project_name)
    create_financial_sheet(wb, summary)
    create_procurement_dna_sheet(wb, procurement_dna)
    existing_titles = {ws.title for ws in wb.worksheets}
    create_scenario_sheet(wb, scenario, safe_sheet_title("Scenariu selectat", "Scenariu selectat", existing_titles))

    assert_no_formulas(wb)
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
