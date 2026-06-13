from __future__ import annotations

import io
import re
import zipfile
from copy import deepcopy
from pathlib import Path
from typing import Any, Dict, Iterable, Tuple
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import DataBarRule


NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "c16r2": "http://schemas.microsoft.com/office/drawing/2015/06/chart",
}
ET.register_namespace("", NS["main"])
ET.register_namespace("c", NS["c"])
ET.register_namespace("a", NS["a"])
ET.register_namespace("r", NS["r"])
ET.register_namespace("c16r2", NS["c16r2"])


MONTHS = ["Ian", "Feb", "Mar", "Apr", "Mai", "Iun", "Iul", "Aug", "Sep", "Oct", "Nov", "Dec"]
NAVY = "1F4E78"
GREEN = "4F9B27"


def as_float(value: Any) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def pick_float(*values: Any) -> float:
    for value in values:
        if value is not None and value != "":
            return as_float(value)
    return 0.0


def fmt(value: Any, decimals: int = 2) -> str:
    return f"{as_float(value):,.{decimals}f}".replace(",", " ")


def col_to_number(col: str) -> int:
    total = 0
    for char in col:
        total = total * 26 + ord(char.upper()) - 64
    return total


def split_cell(cell: str) -> Tuple[str, int]:
    match = re.match(r"^([A-Z]+)(\d+)$", cell)
    if not match:
        raise ValueError(f"Adresa celula invalida: {cell}")
    return match.group(1), int(match.group(2))


def row_number(cell: str) -> int:
    return split_cell(cell)[1]


def cell_sort_key(cell: str) -> Tuple[int, int]:
    col, row = split_cell(cell)
    return row, col_to_number(col)


def q(tag: str) -> str:
    return f"{{{NS['main']}}}{tag}"


def cq(tag: str) -> str:
    return f"{{{NS['c']}}}{tag}"


def aq(tag: str) -> str:
    return f"{{{NS['a']}}}{tag}"


def find_or_create_row(sheet: ET.Element, index: int) -> ET.Element:
    sheet_data = sheet.find("main:sheetData", NS)
    if sheet_data is None:
        sheet_data = ET.SubElement(sheet, q("sheetData"))

    for row in sheet_data.findall("main:row", NS):
        if int(row.attrib.get("r", "0")) == index:
            return row

    new_row = ET.Element(q("row"), {"r": str(index)})
    rows = sheet_data.findall("main:row", NS)
    inserted = False
    for pos, row in enumerate(rows):
        if int(row.attrib.get("r", "0")) > index:
            sheet_data.insert(pos, new_row)
            inserted = True
            break
    if not inserted:
        sheet_data.append(new_row)
    return new_row


def find_or_create_cell(sheet: ET.Element, ref: str) -> ET.Element:
    col, row_idx = split_cell(ref)
    row = find_or_create_row(sheet, row_idx)
    for cell in row.findall("main:c", NS):
        if cell.attrib.get("r") == ref:
            return cell

    new_cell = ET.Element(q("c"), {"r": ref})
    cells = row.findall("main:c", NS)
    inserted = False
    for pos, cell in enumerate(cells):
        existing_ref = cell.attrib.get("r", "")
        existing_col = split_cell(existing_ref)[0] if existing_ref else ""
        if existing_col and col_to_number(existing_col) > col_to_number(col):
            row.insert(pos, new_cell)
            inserted = True
            break
    if not inserted:
        row.append(new_cell)
    return new_cell


def clear_cell(cell: ET.Element) -> None:
    style = cell.attrib.get("s")
    ref = cell.attrib.get("r")
    cell.attrib.clear()
    if ref:
        cell.attrib["r"] = ref
    if style:
        cell.attrib["s"] = style
    for child in list(cell):
        cell.remove(child)


def set_cell(sheet: ET.Element, ref: str, value: Any) -> None:
    cell = find_or_create_cell(sheet, ref)
    clear_cell(cell)
    if value is None or value == "":
        return
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        ET.SubElement(cell, q("v")).text = f"{float(value):.12g}"
        return

    cell.attrib["t"] = "inlineStr"
    inline = ET.SubElement(cell, q("is"))
    text = ET.SubElement(inline, q("t"))
    text.text = str(value)
    if str(value).strip() != str(value):
        text.attrib["{http://www.w3.org/XML/1998/namespace}space"] = "preserve"


def set_cells(sheet: ET.Element, values: Dict[str, Any]) -> None:
    for ref, value in sorted(values.items(), key=lambda item: cell_sort_key(item[0])):
        set_cell(sheet, ref, value)


def update_chart_title(chart_xml: bytes, title: str) -> bytes:
    root = ET.fromstring(chart_xml)
    title_node = root.find(".//c:chart/c:title", NS)
    if title_node is None:
        return chart_xml
    tx = title_node.find("c:tx", NS)
    if tx is None:
        tx = ET.SubElement(title_node, cq("tx"))
    for child in list(tx):
        tx.remove(child)
    rich = ET.SubElement(tx, cq("rich"))
    ET.SubElement(rich, aq("bodyPr"))
    ET.SubElement(rich, aq("lstStyle"))
    paragraph = ET.SubElement(rich, aq("p"))
    run = ET.SubElement(paragraph, aq("r"))
    ET.SubElement(run, aq("rPr"), {"lang": "ro-RO"})
    ET.SubElement(run, aq("t")).text = title
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def observation(reduction: float, diff_vs_contract: float) -> str:
    if diff_vs_contract < 0:
        return "Impact moderat; peste pret contract"
    if reduction >= 130:
        return "Impact ridicat"
    if reduction >= 100:
        return "Impact bun"
    return "Impact moderat"


def build_cell_values(summary: Dict[str, Any]) -> Dict[str, Any]:
    assumptions = summary.get("assumptions") or {}
    annual = summary.get("annual") or {}
    battery = summary.get("battery") or {}
    interpretation = summary.get("interpretation") or {}
    monthly = summary.get("monthly") or []

    client = assumptions.get("client") or "Client"
    scenario = summary.get("scenarioName") or "-"
    ac_mw = as_float(battery.get("acMw"))
    storage_mwh = as_float(battery.get("storageMwh"))
    contract_price = pick_float(annual.get("contract_price_lei_mwh"), assumptions.get("contractPriceLeiMwh"))
    avg_with_bess = as_float(annual.get("arithmetic_avg_monthly_with_bess_lei_mwh"))
    gross_diff = pick_float(
        annual.get("difference_vs_contract_lei_mwh"),
        annual.get("contract_vs_pzu_with_bess_lei_mwh"),
        contract_price - avg_with_bess,
    )
    supply_component = pick_float(annual.get("supply_component_lei_mwh"), assumptions.get("supplyComponentLeiMwh"))
    final_result = pick_float(annual.get("final_result_lei_mwh"), gross_diff - supply_component)
    eur_rate = as_float(assumptions.get("eurRate")) or 5.1
    investment = as_float(annual.get("investment_eur") or assumptions.get("investmentEur"))

    values: Dict[str, Any] = {
        "A1": (
            f"Sinteza financiara  - Client {client} | capacitate baterie "
            f"{fmt(ac_mw, 3)} MW AC - stocare {fmt(storage_mwh, 3)} MWh"
        ),
        "B4": scenario,
        "B5": as_float(annual.get("consumption_mwh")),
        "B6": f"{fmt(interpretation.get('dayPct'), 1)}% / {fmt(interpretation.get('nightPct'), 1)}%",
        "B7": as_float(annual.get("weighted_price_with_bess_lei_mwh")),
        "B8": contract_price,
        "A9": "Componenta furnizare PZU",
        "B9": supply_component,
        "E4": (
            f"1. Scenariul PZU cu BESS este comparat cu pretul contractual de "
            f"{fmt(contract_price)} lei/MWh."
        ),
        "E6": (
            f"2. Diferenta contract vs PZU cu BESS este de "
            f"{fmt(gross_diff)} lei/MWh inainte de componenta furnizare."
        ),
        "E7": (
            f"3. Componenta furnizare este {fmt(supply_component)} lei/MWh, iar impactul anual indicativ este de aproximativ "
            f"{fmt(annual.get('impact_lei_year'), 0)} lei/an."
        ),
        "A11": "Luna",
        "B11": "Pret contract 2025",
        "C11": "PZU cu BESS",
        "D11": "Pret contract 2025 vs PZU cu BESS",
        "E11": "Observatie",
        "F11": "",
        "G11": "",
        "A25": "Sinteza anuala",
        "B25": "Pret Contract 2025",
        "C25": "Pret PZU Cu BESS 2025",
        "D25": "Reducere Pret Contract 2025 VS PZU CU BESS",
        "E25": "Componenta furnizare PZU (Comision furnizor)",
        "F25": "Rezultate",
        "G25": "Comentariu",
        "B26": "=B8",
        "C26": avg_with_bess,
        "D26": "=B26-C26",
        "E26": "=B9",
        "F26": "=D26-E26",
        "G26": "Medii simple ale celor 12 luni din simulare",
        "F27": "=F26*B5",
        "G27": "Consum total x rezultat net",
        "F29": f"=F27/{eur_rate:.12g}" if eur_rate > 0 else "",
        "F31": investment,
        "F33": '=IF(F29>0,F31/F29,"")',
        "I42": "Luna",
        "J42": "Pret contract 2025",
        "K42": "PZU cu BESS",
        "L42": "",
        "M42": "",
    }

    for index in range(12):
        row = monthly[index] if index < len(monthly) else {}
        excel_row = 12 + index
        month = row.get("month") or MONTHS[index]
        with_bess = as_float(row.get("price_with_bess_lei_mwh"))
        diff = pick_float(row.get("difference_vs_contract_lei_mwh"), contract_price - with_bess)
        values.update(
            {
                f"A{excel_row}": month,
                f"B{excel_row}": "=$B$8",
                f"C{excel_row}": with_bess,
                f"D{excel_row}": f"=B{excel_row}-C{excel_row}",
                f"E{excel_row}": row.get("observation") or observation(diff, diff),
                f"F{excel_row}": "",
                f"G{excel_row}": "",
            }
        )

        helper_row = 43 + index
        values.update(
            {
                f"I{helper_row}": f"=A{excel_row}",
                f"J{helper_row}": f"=B{excel_row}",
                f"K{helper_row}": f"=C{excel_row}",
                f"L{helper_row}": "",
                f"M{helper_row}": "",
            }
        )
        reduction_row = 57 + index
        values.update({f"I{reduction_row}": "", f"J{reduction_row}": ""})

    return values


def configure_price_chart(ws, title: str) -> None:
    values = [as_float(ws["B8"].value)]
    values.extend(as_float(ws[f"C{row}"].value) for row in range(12, 24))
    y_axis_max = max(values or [0]) * 1.15

    if ws._charts:
        chart = ws._charts[0]
        chart.title = title
        if y_axis_max > 0:
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = y_axis_max
        while len(chart.series) > 2:
            del chart.series[-1]
        if len(chart.series) >= 2:
            chart.series[0].tx = SeriesLabel(v="Pret contract 2025")
            chart.series[1].tx = SeriesLabel(v="PZU cu BESS")
            chart.series[0].graphicalProperties.line.solidFill = NAVY
            chart.series[0].graphicalProperties.line.width = 32000
            chart.series[0].marker.symbol = "circle"
            chart.series[0].marker.size = 5
            chart.series[1].graphicalProperties.line.solidFill = GREEN
            chart.series[1].graphicalProperties.line.width = 28000
            chart.series[1].marker.symbol = "circle"
            chart.series[1].marker.size = 5
            return
        ws._charts = []

    chart = LineChart()
    chart.title = title
    chart.y_axis.title = "lei/MWh"
    chart.x_axis.title = "Luna"
    if y_axis_max > 0:
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = y_axis_max
    chart.height = 7
    chart.width = 15
    data = Reference(ws, min_col=10, max_col=11, min_row=42, max_row=54)
    cats = Reference(ws, min_col=9, min_row=43, max_row=54)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    if len(chart.series) >= 2:
        chart.series[0].tx = SeriesLabel(v="Pret contract 2025")
        chart.series[0].graphicalProperties.line.solidFill = NAVY
        chart.series[0].graphicalProperties.line.width = 32000
        chart.series[0].marker.symbol = "circle"
        chart.series[0].marker.size = 5
        chart.series[1].tx = SeriesLabel(v="PZU cu BESS")
        chart.series[1].graphicalProperties.line.solidFill = GREEN
        chart.series[1].graphicalProperties.line.width = 28000
        chart.series[1].marker.symbol = "circle"
        chart.series[1].marker.size = 5
    ws.add_chart(chart, "F11")


def ensure_result_data_bars(ws) -> None:
    target = "D12:D23"
    for cf in ws.conditional_formatting:
        if str(cf.sqref) == target and any(rule.type == "dataBar" for rule in cf.rules):
            return
    ws.conditional_formatting.add(
        target,
        DataBarRule(start_type="min", end_type="max", color="70AD47", showValue=True),
    )


def generate_financial_summary_xlsx(summary: Dict[str, Any], template_path: Path) -> bytes:
    if not template_path.exists():
        raise FileNotFoundError(f"Template sinteza lipsa: {template_path}")

    cell_values = build_cell_values(summary)
    assumptions = summary.get("assumptions") or {}
    client = assumptions.get("client") or "Client"
    scenario = summary.get("scenarioName") or "-"
    chart_title = f"{client} - {scenario} | pret contract 2025 vs PZU cu BESS [lei/MWh]"

    output = io.BytesIO()
    wb = load_workbook(template_path)
    ws = wb.worksheets[0]
    for ref, value in cell_values.items():
        ws[ref] = value

    configure_price_chart(ws, chart_title)
    ensure_result_data_bars(ws)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    wb.save(output)
    return output.getvalue()
